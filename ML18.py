from __future__ import unicode_literals, print_function, division
import pickle
import re
from sklearn.metrics import confusion_matrix
from sklearn.linear_model import LogisticRegression
from sklearn.linear_model import Perceptron
from sklearn.linear_model  import PassiveAggressiveClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.svm import SVC
from sklearn.feature_extraction.text import CountVectorizer
import openpyxl
import numpy as np
import tqdm
import warnings
import random
import multiprocessing as mp

warnings.filterwarnings("ignore")
datafrom = openpyxl.load_workbook('NSCLCPD-1.xlsx')
train = datafrom.get_sheet_by_name("Training")
amountdata = []
AccuracyMLP = []
AccuracyLogistic = []
x_pointless = []
data = []
for t in range(2,2252):
    a=train.cell(row = t, column=12).value
    b=a[0:a.rfind("ictated")-1]
    if (train.cell(row = t,column = 3).value in ("POD","SD","POD/brain")):
        c = 0
        data.append((b,c))
    elif (train.cell(row = t,column = 3).value in ("PR","CR")):
        c = 1
        data.append((b,c))
    else:
        print("loading error")
random.shuffle(data)
avgmlp = []
trainData = [data[j] for j in range(int(len(data)*.9))]
testData = [data[u] for u in range(int(len(data)*.9),(int(len(data))))]
trainText, trainY = [d[0] for d in trainData], [d[1] for d in trainData]
testText, testY = [d[0] for d in testData], [d[1] for d in testData]
min_df = 1
max_features = 15000
countVec = CountVectorizer(ngram_range=(1,3), min_df = min_df, max_features = max_features)
trainX = countVec.fit_transform(trainText)
testX = countVec.transform(testText)
mlp = MLPClassifier(hidden_layer_sizes=(100,), alpha = .0001, batch_size='auto', learning_rate='constant', learning_rate_init=0.001, power_t=0.5,max_iter=2000,shuffle=True, random_state=None, tol=0.0001, momentum=0.9)
mlp.fit(trainX, trainY)
score = mlp.score(testX,testY)
predictions = mlp.predict(testX)
confused_matrix = confusion_matrix(testY,predictions)
print(score)                                         
print(confused_matrix)                                                                                                                            
    # falpos = 0
    # falneg = 0
    # truepos = 0
    # trueneg = 0
    # testPredict = []
    # accactual = []
    # truepostotal = []
    # truenegtotal = []
    # falpostotal = []
    # falnegtotal = []
    # thresholdlog = []
#    for t in range(0,41):
        # for i in range((len(testY))):
        #     a = mlp.predict_proba((testX[i]))
        #     if ((a[0][0]) >= ((.5))):
        #         testPredict.append(0)
        #         #print((a[0][0]))
        #         #print((testPredict[i]),(testY[i]),((mlp.predict_proba(testX[i]))))
        #     else:
        #         testPredict.append(1)
        #
        # for i in range(len(testY)):
        #     a = testPredict[i]
        #     b = int(a)
        #     c = testY[i]
        #     if (b != c):
        #         if (b == 1):
        #             # print((i,"false positive"))
        #             falpos += 1
        #         else:
        #             # print((i,"false negative"))
        #             falneg +=1
        #     else:
        #         if (b == 1):
        #             # print((i, "true positive"))
        #             truepos += 1
        #         else:
        #             # print((i, "true negative"))
        #             trueneg += 1
        # falposrte = (falpos/(falpos+trueneg))
        # falnegrte = (falneg/(falneg+truepos))
        # trueposrte = (truepos/(falneg+truepos))
        # truenegrte = (trueneg/(falpos+trueneg))
        # #print(trueposrte)
        # #print(("threshold:",(.05*t),"false positive rate ",falposrte,"false negative rate ",falnegrte,"true positive rate ",trueposrte,"true negative rate ",truenegrte))
        # mlp_score = mlp.score(testX, testY)
        # truepostotal.append(trueposrte)
        # truenegtotal.append(truenegrte)
        # falpostotal.append(falposrte)
        # falnegtotal.append(falnegrte)
        # thresholdlog.append((.025*t))
        # falpos = 0
        # falneg = 0
        # truepos = 0
        # trueneg = 0
        # testPredict = []
        # accactual = []
    # for t in range(1,100):
    #     for i in range((len(testY))):
    #         a = mlp.predict_proba((testX[i]))
    #         if ((a[0][0]) >= (.9+(.001*t))):
    #             testPredict.append(0)
    #             #print((a[0][0]))
    #             #print((testPredict[i]),(testY[i]),((mlp.predict_proba(testX[i]))))
    #         else:
    #             testPredict.append(1)
    #
    #     for i in range(len(testY)):
    #         a = testPredict[i]
    #         b = int(a)
    #         c = testY[i]
    #         if (b != c):
    #             if (b == 1):
    #                 # print((i,"false positive"))
    #                 falpos += 1
    #             else:
    #                 # print((i,"false negative"))
    #                 falneg +=1
    #         else:
    #             if (b == 1):
    #                 # print((i, "true positive"))
    #                 truepos += 1
    #             else:
    #                 # print((i, "true negative"))
    #                 trueneg += 1
    #     falposrte = (falpos/(falpos+trueneg))
    #     falnegrte = (falneg/(falneg+truepos))
    #     trueposrte = (truepos/(falneg+truepos))
    #     truenegrte = (trueneg/(falpos+trueneg))
    #     truepostotal.append(trueposrte)
    #     truenegtotal.append(truenegrte)
    #     falpostotal.append(falposrte)
    #     falnegtotal.append(falnegrte)
    #     thresholdlog.append((.9+(.001*t)))
    #     #print(("threshold:",((.9+(.001*t))),"false positive rate ",falposrte,"false negative rate ",falnegrte,"true positive rate ",trueposrte,"true negative rate ",truenegrte))
    #     falpos = 0
    #     falneg = 0
    #     truepos = 0
    #     trueneg = 0
    #     testPredict = []
    #     accactual = []
    # print("a = ", falpostotal)
    # print("b =", falnegtotal)
    # print("c =", truepostotal)
    # print("d =", truenegtotal)
    # print("e =", thresholdlog)
